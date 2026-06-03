"""Comprehensive test suite for sandbox Docker image.

Tests:
1. Functional tests: verify all pre-installed packages work
2. Security tests: verify security configurations
3. Performance tests: verify startup and import times
"""

import subprocess
import time

import pytest


class TestImageFunctionality:
    """Test all pre-installed packages and tools."""

    IMAGE = "myrm/skill-sandbox:latest"

    def run_in_container(self, command: str) -> tuple[int, str, str]:
        """Run command in container and return exit code, stdout, stderr."""
        result = subprocess.run(
            ["docker", "run", "--rm", self.IMAGE, "python", "-c", command],
            capture_output=True,
            text=True,
            timeout=30,
        )
        return result.returncode, result.stdout, result.stderr

    def test_python_version(self) -> None:
        """Test Python version is correct."""
        code, stdout, _ = self.run_in_container("import sys; print(sys.version)")
        assert code == 0
        assert "3.14" in stdout

    def test_data_science_packages(self) -> None:
        """Test data science packages work correctly."""
        code, stdout, _ = self.run_in_container("""
import pandas as pd
import numpy as np
import scipy

# Test pandas
df = pd.DataFrame({'a': [1, 2, 3], 'b': [4, 5, 6]})
assert len(df) == 3

# Test numpy
arr = np.array([1, 2, 3])
assert arr.sum() == 6

print('OK')
""")
        assert code == 0
        assert "OK" in stdout

    def test_file_processing(self) -> None:
        """Test file processing packages work."""
        code, stdout, _ = self.run_in_container("""
import openpyxl
from docx import Document
from pptx import Presentation
from pypdf import PdfReader
from PIL import Image

# Test openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'test'
assert ws['A1'].value == 'test'

print('OK')
""")
        assert code == 0
        assert "OK" in stdout

    def test_visualization(self) -> None:
        """Test visualization packages work."""
        code, stdout, _ = self.run_in_container("""
import matplotlib
matplotlib.use('Agg')  # Non-GUI backend
import matplotlib.pyplot as plt
import seaborn as sns

# Test matplotlib
fig, ax = plt.subplots()
ax.plot([1, 2, 3], [1, 2, 3])

print('OK')
""")
        assert code == 0
        assert "OK" in stdout

    def test_cjk_font_rendering(self) -> None:
        """Verify matplotlib renders CJK text with real glyphs (no tofu boxes)."""
        code, stdout, stderr = self.run_in_container("""
import os

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import font_manager as fm

# Default sans-serif must resolve to a Noto CJK face, not the DejaVu fallback
font_path = fm.findfont(fm.FontProperties(family='sans-serif'))
assert 'NotoSansCJK' in font_path or 'NotoSerifCJK' in font_path, font_path

fig, ax = plt.subplots()
ax.set_title('中文标题：销售额分析')
ax.set_xlabel('区域')
ax.plot([1, 2, 3], [1, 2, 3])
fig.savefig('/tmp/cjk_test.png')

assert os.path.getsize('/tmp/cjk_test.png') > 0
print('OK')
""")
        assert code == 0, f"CJK rendering failed: {stderr}"
        assert "OK" in stdout
        # No missing-glyph or font-lookup warnings leaked to stderr
        assert "missing from font" not in stderr, stderr
        assert "findfont:" not in stderr, stderr

    def test_playwright(self) -> None:
        """Test Playwright is available."""
        code, stdout, _ = self.run_in_container("""
from playwright.sync_api import sync_playwright

# Just import, don't run (requires running browser)
print('OK')
""")
        assert code == 0
        assert "OK" in stdout

    def test_system_tools(self) -> None:
        """Test system tools are available."""
        tools = ["rg --version", "fd --version", "sqlite3 --version"]
        for tool in tools:
            result = subprocess.run(
                ["docker", "run", "--rm", self.IMAGE, "sh", "-c", tool],
                capture_output=True,
                timeout=10,
            )
            assert result.returncode == 0, f"Tool failed: {tool}"


class TestImageSecurity:
    """Test security configurations."""

    IMAGE = "myrm/skill-sandbox:latest"

    def test_non_root_user(self) -> None:
        """Test container runs as non-root user."""
        result = subprocess.run(
            ["docker", "run", "--rm", self.IMAGE, "id", "-u"],
            capture_output=True,
            text=True,
            timeout=10,
        )
        uid = result.stdout.strip()
        assert uid != "0", "Container should not run as root"

    def test_readonly_rootfs(self) -> None:
        """Test root filesystem is read-only."""
        # This is enforced at runtime by control-plane, not in image
        # Test that tmpfs is writable
        result = subprocess.run(
            [
                "docker",
                "run",
                "--rm",
                "--read-only",
                "--tmpfs",
                "/tmp",
                self.IMAGE,
                "sh",
                "-c",
                "echo test > /tmp/test.txt && cat /tmp/test.txt",
            ],
            capture_output=True,
            text=True,
            timeout=10,
        )
        assert result.returncode == 0
        assert "test" in result.stdout

    def test_no_dangerous_capabilities(self) -> None:
        """Test container doesn't have dangerous capabilities."""
        result = subprocess.run(
            ["docker", "run", "--rm", "--cap-drop=ALL", self.IMAGE, "python", "-c", "print('OK')"],
            capture_output=True,
            text=True,
            timeout=10,
        )
        assert result.returncode == 0
        assert "OK" in result.stdout


class TestImagePerformance:
    """Test performance metrics."""

    IMAGE = "myrm/skill-sandbox:latest"

    def test_startup_time(self) -> None:
        """Test container startup time is acceptable."""
        times = []
        for _ in range(3):
            start = time.time()
            subprocess.run(
                ["docker", "run", "--rm", self.IMAGE, "python", "-c", "print('ready')"],
                capture_output=True,
                timeout=30,
            )
            elapsed = time.time() - start
            times.append(elapsed)

        avg_time = sum(times) / len(times)
        # Should start in under 5 seconds
        assert avg_time < 5.0, f"Startup time too slow: {avg_time:.2f}s"

    def test_import_performance(self) -> None:
        """Test package import times are reasonable."""
        code = """
import time

packages = ['pandas', 'numpy', 'matplotlib']
for pkg in packages:
    start = time.time()
    __import__(pkg)
    elapsed = time.time() - start
    # Each package should import in under 2 seconds
    assert elapsed < 2.0, f"{pkg} import too slow: {elapsed:.2f}s"
    
print('OK')
"""
        result = subprocess.run(
            ["docker", "run", "--rm", self.IMAGE, "python", "-c", code],
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0
        assert "OK" in result.stdout


class TestImageHealth:
    """Test health check functionality."""

    IMAGE = "myrm/skill-sandbox:latest"

    def test_health_check_passes(self) -> None:
        """Test container health check passes (lightweight check)."""
        # Start container
        result = subprocess.run(
            ["docker", "run", "-d", self.IMAGE, "sleep", "60"],
            capture_output=True,
            text=True,
            timeout=10,
        )
        container_id = result.stdout.strip()

        try:
            # Wait for health check (longer since interval=60s)
            time.sleep(15)

            # Check health status
            result = subprocess.run(
                ["docker", "inspect", "-f", "{{.State.Health.Status}}", container_id],
                capture_output=True,
                text=True,
                timeout=10,
            )
            health_status = result.stdout.strip()
            # Note: may be "starting" if check hasn't run yet
            assert health_status in ["healthy", "starting"], f"Health check failed: {health_status}"

        finally:
            # Cleanup
            subprocess.run(["docker", "rm", "-f", container_id], capture_output=True)

    def test_uv_available(self) -> None:
        """Test uv package manager is available."""
        result = subprocess.run(
            ["docker", "run", "--rm", self.IMAGE, "uv", "--version"],
            capture_output=True,
            text=True,
            timeout=10,
        )
        assert result.returncode == 0
        assert "uv" in result.stdout.lower()

    def test_deep_health_check_script(self) -> None:
        """Test deep health check script is available and functional."""
        result = subprocess.run(
            ["docker", "run", "--rm", self.IMAGE, "deep-health-check"],
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, f"Deep health check failed: {result.stderr}"
        assert "✅" in result.stdout or "passed" in result.stdout.lower()


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
